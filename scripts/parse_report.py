#!/usr/bin/env python3
"""
Разбор ежемесячного отчёта по продажам (xlsx) в чистый JSON для дашборда.

Отчёт ведётся вручную, поэтому парсер занимается в основном нормализацией:
даты в пяти разных форматах, числа с запятой, имена менеджеров кириллицей и
латиницей, направления и операторы в произвольном регистре и на трёх языках.

Запуск:  python3 scripts/parse_report.py <файл.xlsx> [-o data/deals.json]
"""

import argparse
import datetime as dt
import json
import re
import sys
import unicodedata
from collections import Counter

import openpyxl

MONTHS = {"январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
          "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12}

# Раскладка колонок. В листе «май» нет колонки «Брон санаси», поэтому смещение.
LAYOUT_WITH_BOOKING = dict(no=0, booked=1, client=2, pax=3, operator=4, depart=5,
                           ret=6, direction=7, gross=8, net=9, profit=10,
                           received=11, client_debt=12, paid_operator=13,
                           hotel=14, transport=15, manager=16)
LAYOUT_NO_BOOKING = dict(no=0, booked=None, client=1, pax=2, operator=3, depart=4,
                         ret=5, direction=6, gross=7, net=8, profit=9,
                         received=10, client_debt=11, paid_operator=12,
                         hotel=13, transport=14, manager=15)

MANAGERS = {
    "сарвиноз": "Сарвиноз", "sarvinoz": "Сарвиноз",
    "муслимжон": "Муслимжон", "muslimjon": "Муслимжон", "муслимжан": "Муслимжон",
    "азиза": "Азиза", "aziza": "Азиза",
    "абдулкарим": "Абдулкарим", "abdulkarim": "Абдулкарим",
    "мадина": "Мадина", "madina": "Мадина",
    "шахноза": "Шахноза", "shahnoza": "Шахноза", "shaxnoza": "Шахноза",
}

# Направления. Ключ — нормализованная подстрока, значение — каноническое имя.
# Порядок важен: проверяется вхождение, первое совпадение выигрывает.
DIRECTIONS = [
    ("шарм", "Египет · Шарм-эль-Шейх"), ("sharm", "Египет · Шарм-эль-Шейх"),
    ("хургада", "Египет · Хургада"), ("hurgada", "Египет · Хургада"),
    ("батуми", "Грузия · Батуми"), ("batumi", "Грузия · Батуми"),
    ("трабзон", "Грузия · Батуми"),
    ("грузия", "Грузия"), ("gruziya", "Грузия"),
    ("вьетнам", "Вьетнам"), ("vietnam", "Вьетнам"),
    ("таиланд", "Таиланд"), ("thailand", "Таиланд"), ("tailand", "Таиланд"),
    ("анталья", "Турция"), ("antalya", "Турция"), ("бодрум", "Турция"),
    ("bodrum", "Турция"), ("турция", "Турция"), ("turkiya", "Турция"),
    ("turkey", "Турция"),
    ("дубай", "ОАЭ"), ("dubai", "ОАЭ"), ("оаэ", "ОАЭ"), ("emirat", "ОАЭ"),
    ("хайнань", "Китай"), ("haynan", "Китай"), ("xitoy", "Китай"),
    ("hitoy", "Китай"), ("китай", "Китай"), ("china", "Китай"),
    ("бали", "Индонезия · Бали"), ("bali", "Индонезия · Бали"),
    ("малайзия", "Малайзия"), ("malaysia", "Малайзия"),
    ("сингапур", "Сингапур"), ("singapore", "Сингапур"),
    ("чехия", "Европа · Чехия"), ("czech", "Европа · Чехия"),
    ("петербург", "Россия"), ("москва", "Россия"),
    ("умра", "Умра"), ("umra", "Умра"),
    ("виза", "Визовые услуги"), ("viza", "Визовые услуги"),
    ("нафталан", "Азербайджан"), ("naftalan", "Азербайджан"),
    ("иссык", "Кыргызстан"), ("issyk", "Кыргызстан"),
    ("узбекистан", "Узбекистан"), ("uzbekistan", "Узбекистан"),
]

OPERATORS = {
    "turon": "Turon", "easybooking": "EasyBooking", "pegas": "Pegas",
    "prestige": "Prestige", "kompas": "Kompas", "asialuxe": "Asialuxe",
    "flykhiva": "FlyKhiva", "fun & sun": "Fun & Sun", "fun&sun": "Fun & Sun",
    "crystal bay": "Crystal Bay", "kazunion": "Kazunion", "velvek": "Velvek",
    "centrum": "Centrum", "centrum holidays": "Centrum", "batik": "Batik",
    "izel": "Izel", "selfie": "Selfie", "etihad": "Etihad",
    "vipturkey": "VIPTurkey", "ratehawk": "Ratehawk",
    "individual": "Индивидуальный", "indvidual": "Индивидуальный",
    "bilet": "Авиабилет", "avia": "Авиабилет", "malva": "Malva",
}


def norm(s):
    """Нижний регистр, схлопнутые пробелы, ё→е."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s)).strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def parse_number(v):
    """'4548,5' → 4548.5. Пустое и мусор → None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in {"-", "."}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(v, year=2026):
    """
    Даты в отчёте лежат в трёх видах:
      • '12.05.2026'      — нормальная строка ДД.ММ.ГГГГ
      • '17/07'           — ДД/ММ без года
      • datetime(2026,11,7) — Excel прочитал ДД/ММ как ММ/ДД и перевернул

    Третий случай — самый коварный: ячейка выглядит как дата, но день и
    месяц в ней переставлены. Excel переворачивает только когда день ≤ 12
    (иначе он не смог бы принять его за месяц и оставил бы строкой), значит
    у всех datetime-ячеек день и месяц нужно вернуть на место.
    """
    if v is None:
        return None
    if isinstance(v, dt.datetime) or isinstance(v, dt.date):
        d = v.date() if isinstance(v, dt.datetime) else v
        if d.day <= 12:
            try:
                return dt.date(d.year, d.day, d.month).isoformat()
            except ValueError:
                pass
        return d.isoformat()

    s = str(v).strip()
    if not s or s == "-":
        return None
    parts = re.split(r"[./\-\s]+", s)
    parts = [p for p in parts if p]
    try:
        if len(parts) >= 3:
            day, month, yr = int(parts[0]), int(parts[1]), int(parts[2])
            if yr < 100:
                yr += 2000
            return dt.date(yr, month, day).isoformat()
        if len(parts) == 2:
            return dt.date(year, int(parts[1]), int(parts[0])).isoformat()
    except (ValueError, IndexError):
        return None
    return None


def map_direction(v):
    n = norm(v)
    if not n or n == "-":
        return None, None
    for key, canon in DIRECTIONS:
        if key in n:
            return canon, None
    return v.strip() if isinstance(v, str) else str(v), n  # вернём как есть + флаг


def map_operator(v):
    n = norm(v)
    if not n or n == "-":
        return None, None
    for key, canon in OPERATORS.items():
        if key in n:
            return canon, None
    return (v.strip() if isinstance(v, str) else str(v)), n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("-o", "--out", default="data/deals.json")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    deals, issues = [], []
    unmapped_dir, unmapped_op = Counter(), Counter()
    ad_spend = {}

    for ws in wb.worksheets:
        sheet = norm(ws.title)
        month = MONTHS.get(sheet)
        if month is None:
            issues.append(f"лист «{ws.title}»: не удалось определить месяц, пропущен")
            continue
        period = f"2026-{month:02d}"

        header = [norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        L = LAYOUT_WITH_BOOKING if any("брон" in h for h in header[:3]) else LAYOUT_NO_BOOKING

        # Расходы на таргетированную рекламу — подписаны сбоку свободным текстом.
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=17, max_col=24, values_only=True):
            for i, cell in enumerate(row):
                n = norm(cell)
                if "таргет" in n or n == "target":
                    m = re.search(r"(\d[\d\s.,]*)", n)
                    if m:
                        ad_spend[period] = parse_number(m.group(1))
                    elif i + 1 < len(row):
                        val = parse_number(row[i + 1])
                        if val:
                            ad_spend[period] = val

        for excel_row, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r[L["no"]] is None:
                continue
            manager_raw = r[L["manager"]]
            manager = MANAGERS.get(norm(manager_raw))
            gross = parse_number(r[L["gross"]])
            net = parse_number(r[L["net"]])

            if manager is None or gross is None:
                if manager_raw or gross:
                    issues.append(
                        f"{ws.title}, строка {excel_row}: "
                        f"{'неизвестный менеджер «%s»' % manager_raw if manager is None else 'нет суммы'}"
                    )
                continue

            direction, bad_dir = map_direction(r[L["direction"]])
            operator, bad_op = map_operator(r[L["operator"]])
            if bad_dir:
                unmapped_dir[bad_dir] += 1
            if bad_op:
                unmapped_op[bad_op] += 1

            # Прибыль. Считаем сами из брутто и нетто — это единственный
            # способ получить цифру, которая не разъезжается с суммами.
            # Если нетто не заполнено, прибыль неизвестна: подставлять сюда
            # брутто нельзя, иначе сделка без себестоимости выглядит как
            # сверхприбыльная и ломает весь рейтинг менеджеров.
            profit_stated = parse_number(r[L["profit"]])
            gaps = []
            if net is None:
                gaps.append("не заполнено нетто")
                profit = profit_stated
                if profit is None:
                    gaps.append("прибыль не посчитана")
            else:
                profit = round(gross - net, 2)
                if profit_stated is not None and abs(profit - profit_stated) > 1:
                    gaps.append(f"в отчёте прибыль {profit_stated:.0f}, брутто−нетто = {profit:.0f}")
                    issues.append(f"{ws.title}, строка {excel_row}: "
                                  f"прибыль в отчёте {profit_stated:.0f}, "
                                  f"а брутто−нетто = {profit:.0f}")
            if gaps and net is None:
                issues.append(f"{ws.title}, строка {excel_row}: " + ", ".join(gaps))

            net_v = net if net is not None else 0.0
            paid_op = parse_number(r[L["paid_operator"]]) or 0.0

            deals.append({
                "period": period,
                "row": excel_row,
                "manager": manager,
                "client": (str(r[L["client"]]).strip() if r[L["client"]] else None),
                "pax": int(parse_number(r[L["pax"]]) or 1),
                "operator": operator,
                "direction": direction,
                "hotel": (str(r[L["hotel"]]).strip() if r[L["hotel"]] else None),
                "booked_at": parse_date(r[L["booked"]], 2026) if L["booked"] is not None else None,
                "depart_at": parse_date(r[L["depart"]], 2026),
                "return_at": parse_date(r[L["ret"]], 2026),
                "gross": round(gross, 2),
                "net": (round(net, 2) if net is not None else None),
                "profit": (round(profit, 2) if profit is not None else None),
                "gaps": gaps,
                "received": round(parse_number(r[L["received"]]) or 0.0, 2),
                "client_debt": round(parse_number(r[L["client_debt"]]) or 0.0, 2),
                "paid_operator": round(paid_op, 2),
                # Сколько агентство ещё должно туроператору. В отчёте этой
                # колонки нет, но она выводится и это самый горящий показатель.
                "operator_debt": round(max(net_v - paid_op, 0), 2) if net is not None else None,
            })

    for v, n in unmapped_dir.items():
        issues.append(f"направление «{v}» ({n} шт.) не нашлось в справочнике — оставлено как есть")
    for v, n in unmapped_op.items():
        issues.append(f"туроператор «{v}» ({n} шт.) не нашёлся в справочнике — оставлен как есть")

    payload = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source": args.xlsx.split("/")[-1],
        "commission_rate": 0.5,       # менеджер получает 50% от прибыли по сделке
        "ad_spend_split": 8,          # расход на рекламу делится на 8 долей
        "ad_spend": ad_spend,
        "deals": deals,
        "issues": issues,
    }

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    print(f"Сделок разобрано: {len(deals)}")
    print(f"Расходы на рекламу по месяцам: {ad_spend}")
    print(f"Замечаний к данным: {len(issues)}")
    for i in issues[:40]:
        print("  •", i)
    if len(issues) > 40:
        print(f"  … ещё {len(issues) - 40}")


if __name__ == "__main__":
    sys.exit(main())
